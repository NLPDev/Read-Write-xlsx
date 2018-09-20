

from openpyxl import Workbook
from openpyxl.styles import Alignment


import xlrd
import os


loc = ("aa.xlsx")

wbr = xlrd.open_workbook(loc)


cnt=0
for wsh in wbr.sheet_names():

	sheet=wbr.sheet_by_index(cnt)

	try:
		# Create target Directory
		os.mkdir(wsh)

	except FileExistsError:
		mm=0

	nr = sheet.nrows
	nc = sheet.ncols

	for i in range(nr - 1):
		st = sheet.cell_value(i + 1, 0)

		wb = Workbook()
		sh = wb.active
		sh.title = 'Sheet1'
		sh.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc - 1)
		# sh.merge_cells('A1:C1')
		c2 = sh.cell(row=1, column=1)
		c2.value = "Table1"
		c2.alignment = Alignment(horizontal='center', vertical='center')

		for j in range(nc - 1):
			c2 = sh.cell(row=2, column=j + 1)
			c2.value = sheet.cell_value(0, j + 1)
			st_data = sheet.cell_value(i + 1, j + 1)
			st_split = st_data.split(",")
			k = 0
			for kk in st_split:
				c2 = sh.cell(row=k + 3, column=j + 1)
				c2.value = kk
				k = k + 1


		st = st + ".xlsx"
		ss = "./"+wsh+"/" + st
		wb.save(ss)

